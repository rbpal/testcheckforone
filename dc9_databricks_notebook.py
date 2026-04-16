# Databricks notebook source
# DC-9 Automated Audit Control Testing — Federated Hermes Limited
# Ported from local src/ pipeline. Only paths and credentials change.
# All Python logic is identical to the local pipeline.

# COMMAND ----------

# MAGIC %md
# MAGIC # DC-9 Automated Audit Control Testing
# MAGIC **Federated Hermes Limited — Revenue Detective Control**
# MAGIC
# MAGIC Three-layer architecture:
# MAGIC - **Layer 1:** Deterministic Python checks (no LLM)
# MAGIC - **Layer 2:** Direct Azure OpenAI call → PASS narrative
# MAGIC - **Layer 3:** LangGraph ReAct agent → EXCEPTION investigation + narrative
# MAGIC
# MAGIC Expected: 11 PASS + 1 EXCEPTION (Q3 Attribute D — billing rate changed 28.5bps → 30bps)

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 1 — Install packages
# MAGIC openpyxl and pandas are pre-installed on Databricks Runtime.
# MAGIC langchain, langchain-openai and langgraph are not.

# COMMAND ----------

# Task 8.1 — pip installs
# %pip install openpyxl langchain langchain-openai langchain-core langgraph python-dotenv

# NOTE: Uncomment the line above when running on Databricks.
# openpyxl is NOT pre-installed on Databricks Runtime 16.4 LTS — must be installed explicitly.
# Commented here so this file can be linted/imported locally without error.

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 2 — Configuration (replaces src/config.py)
# MAGIC File paths adapted for Databricks workspace. All cell addresses unchanged.

# COMMAND ----------

# ─────────────────────────────────────────────────────────────────────────
# Task 8.2 — Databricks file paths (Unity Catalog Volume)
# ─────────────────────────────────────────────────────────────────────────
# Inputs/outputs live in an ADLS container (fhidevgtohackadls / team5)
# exposed to Databricks as a Unity Catalog Volume. The Volume root is
# mapped DIRECTLY to the container root — there is no intermediate
# `adls_storage/` subfolder inside team5. The Volume name `adls_storage`
# appears in the Databricks path but NOT in the ADLS path.
#
# Files dropped into ADLS via Azure Storage Explorer appear automatically
# under the corresponding Volume path within 1–10 seconds (ADLS eventual
# consistency window — the guard clause below handles it).
#
#           ADLS path                          Databricks Volume path
#           ─────────                          ──────────────────────
#
#           team5/                             /Volumes/dev-gto-hack-catalog/
#            │                                   agentic_audit_architects/
#            │   ◄══════ SAME BYTES ═══════►     adls_storage/
#            │
#            ├── testFileCopy                   ├── testFileCopy
#            └── data/                          └── data/
#                ├── toc/                           ├── toc/
#                │   └── _TOC.xlsx                  │   └── _TOC.xlsx
#                ├── input/                         ├── input/
#                └── output/                        └── output/
#
# Everything below team5/ on the ADLS side = everything below
# adls_storage/ on the Databricks side. One byte, two names.
#
# Row-by-row mapping:
#   team5/data/input/   ←→  /Volumes/dev-gto-hack-catalog/agentic_audit_architects/adls_storage/data/input/
#   team5/data/toc/     ←→  /Volumes/dev-gto-hack-catalog/agentic_audit_architects/adls_storage/data/toc/
#   team5/data/output/  ←→  /Volumes/dev-gto-hack-catalog/agentic_audit_architects/adls_storage/data/output/
#
# File Arrival Trigger watches data/input/ and fires the job when Q1 + Q3
# workbooks land. The guard clause in Cell 8b waits for both files before
# proceeding, so the trigger cannot run the pipeline half-populated.
UC_CATALOG = "dev-gto-hack-catalog"
UC_SCHEMA  = "agentic_audit_architects"
UC_VOLUME  = "adls_storage"
VOLUME_BASE = f"/Volumes/{UC_CATALOG}/{UC_SCHEMA}/{UC_VOLUME}"

DATA_INPUT_DIR  = f"{VOLUME_BASE}/data/input"
DATA_TOC_DIR    = f"{VOLUME_BASE}/data/toc"
DATA_OUTPUT_DIR = f"{VOLUME_BASE}/data/output"

# Source workbooks — replaced every quarter. These are the AUM Billing
# workbooks produced by the business; the pipeline reads them as inputs.
Q1_FILENAME  = "Anonymised_DC-8_1_Q1_2025.xlsx"
Q3_FILENAME  = "Anonymised_DC_8_7_Q3_2025.xlsx"

# Output template — read-only, structural shell. The TOC template is NOT a
# source workbook; it is the blank DC-9 Test of Controls spreadsheet that the
# pipeline populates with checkmarks, exception marks and narrative evidence.
# Same template is reused across quarters. Lives in data/toc/ (sibling of
# data/input/ and data/output/) to separate it from quarterly source
# workbooks and signal its read-only, template-shell role.
TOC_FILENAME = "4_06_2025_Revenue _TOC.xlsx"          # note: space before _TOC

from datetime import datetime
from zoneinfo import ZoneInfo
_now_est = datetime.now(ZoneInfo("America/New_York"))
OUTPUT_FILENAME = _now_est.strftime("%Y-%m-%d_%H-%M_%Z_Revenue_TOC_COMPLETED.xlsx")

Q1_FILE     = f"{DATA_INPUT_DIR}/{Q1_FILENAME}"
Q3_FILE     = f"{DATA_INPUT_DIR}/{Q3_FILENAME}"
TOC_FILE    = f"{DATA_TOC_DIR}/{TOC_FILENAME}"
OUTPUT_FILE = f"{DATA_OUTPUT_DIR}/{OUTPUT_FILENAME}"

# Tab names (unchanged from local)
CHECKLIST_TAB      = "Checklist"
BILLING_SUMMARY_TAB = "Billing Summary"
TOC_DC9_TAB        = "DC-9"

# Checklist tab cell references
ATTRIBUTE_ROW_MAP = {"A": 25, "B": 26, "C": 28, "D": 29, "E": 30, "F": 31}
CHECKLIST_CHECKED_BY_COL    = "F"
CHECKLIST_REVIEWED_BY_COL   = "G"
CHECKLIST_ANNOTATION_COL    = "H"
CHECKLIST_IA_ANNOTATION_COL = "I"
CHECKLIST_APPROVAL_ROW_START = 19
CHECKLIST_APPROVAL_ROW_END   = 31
CHECKLIST_PREPARER_NAME_CELL = "C34"
CHECKLIST_PREPARER_DATE_CELL = "G34"
CHECKLIST_REVIEWER_NAME_CELL = "C37"
CHECKLIST_REVIEWER_DATE_CELL = "G37"
CHECKLIST_COMMENTS_COL       = "B"
CHECKLIST_COMMENTS_ROW_START = 40
CHECKLIST_COMMENTS_ROW_END   = 55

# Billing Summary tab cell references
BS_BILLING_RATE_LABEL   = "J33"
BS_BILLING_RATE_DISPLAY = "P17"
BS_BILLING_RATE_DECIMAL = "Q38"
BS_UK_GROSS             = "H19"
BS_UK_LESS_CASH         = "H20"
BS_UK_NET               = "J21"
BS_INTL_GROSS           = "H23"
BS_INTL_LESS_CASH       = "H24"
BS_INTL_NET             = "J25"
BS_HRESDF_GROSS         = "H27"
BS_HRESDF_LESS_INCOME   = "H28"
BS_HRESDF_NET           = "J29"
BS_TOTAL_ASSET_VALUE    = "J31"
BS_QUARTERLY_BILLINGS   = "J35"
BS_BILLING_CHECK        = "N49"
BS_DISPOSAL_ROW_START   = 43
BS_DISPOSAL_ROW_END     = 52
BS_DISPOSAL_NAME_COL    = "B"
BS_DISPOSAL_DAYS_COL    = "G"
BS_DISPOSAL_BILLING_COL = "J"
BS_REBATE_ROW_START     = 55
BS_REBATE_ROW_END       = 76
BS_REBATE_NAME_COL      = "B"
BS_REBATE_AMOUNT_COL    = "J"
BS_REBATE_RATE_TYPE_COL = "L"
FORMULA_CELLS           = ["J35", "J31", "J21", "J25", "J29", "N49", "Q38"]

# TOC DC-9 tab cell references
TOC_ATTR_ROW_MAP = {"A": 37, "B": 38, "C": 39, "D": 40, "E": 41, "F": 42}
TOC_Q1_RESULTS_ROW     = 47
TOC_Q3_RESULTS_ROW     = 48
TOC_CHECKMARK_COLS     = {"A": "H", "B": "I", "C": "J", "D": "K", "E": "L", "F": "M"}
TOC_Q1_NARRATIVE_COL   = "P"
TOC_Q3_NARRATIVE_COL   = "Q"
TOC_NARRATIVE_HEADER_ROW   = 36
TOC_NARRATIVE_ROW_START    = 37
TOC_NARRATIVE_ROW_END      = 42
TOC_Q1_NARRATIVE_HEADER    = "Q1 2025 (DC-8.1) narrative evidence"
TOC_Q3_NARRATIVE_HEADER    = "Q3 2025 (DC-8.7) narrative evidence"
TOC_EXCEPTIONS_NOTED_CELL  = "H58"
CHECKMARK_VALUE    = "a"
CHECKMARK_FONT_NAME = "Webdings"
CHECKMARK_FONT_SIZE = 11
EXCEPTION_VALUE    = "X"
EXCEPTION_FONT_NAME = "Arial"
EXCEPTION_FONT_SIZE = 9
FLOAT_TOLERANCE             = 1.0
BILLING_MOVEMENT_THRESHOLD  = 0.05

print("Configuration loaded.")
print(f"  Q1  : {Q1_FILE}")
print(f"  Q3  : {Q3_FILE}")
print(f"  TOC : {TOC_FILE}")
print(f"  Out : {OUTPUT_FILE}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 3 — Credentials (replaces .env)
# MAGIC Uses Databricks secrets instead of python-dotenv.

# COMMAND ----------

# Task 8.3 — Databricks secrets for Azure OpenAI credentials
import os

# On Databricks: uncomment these four lines and remove the os.environ fallbacks.
# AZURE_OPENAI_ENDPOINT        = dbutils.secrets.get(scope="azure-openai", key="endpoint")
# AZURE_OPENAI_API_KEY         = dbutils.secrets.get(scope="azure-openai", key="api-key")
# AZURE_OPENAI_DEPLOYMENT_NAME = dbutils.secrets.get(scope="azure-openai", key="deployment-name")
# AZURE_OPENAI_API_VERSION     = dbutils.secrets.get(scope="azure-openai", key="api-version")

# Local fallback (reads from environment / .env if already loaded):
AZURE_OPENAI_ENDPOINT        = os.getenv("AZURE_OPENAI_ENDPOINT")
AZURE_OPENAI_API_KEY         = os.getenv("AZURE_OPENAI_API_KEY")
AZURE_OPENAI_DEPLOYMENT_NAME = os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME")
AZURE_OPENAI_API_VERSION     = os.getenv("AZURE_OPENAI_API_VERSION")

# Override environment so LangChain picks them up automatically
os.environ["AZURE_OPENAI_ENDPOINT"]        = AZURE_OPENAI_ENDPOINT or ""
os.environ["AZURE_OPENAI_API_KEY"]         = AZURE_OPENAI_API_KEY or ""
os.environ["AZURE_OPENAI_DEPLOYMENT_NAME"] = AZURE_OPENAI_DEPLOYMENT_NAME or ""
os.environ["AZURE_OPENAI_API_VERSION"]     = AZURE_OPENAI_API_VERSION or ""

endpoint_display = (AZURE_OPENAI_ENDPOINT or "NOT SET").replace(
    AZURE_OPENAI_API_KEY or "x", "***"
) if AZURE_OPENAI_ENDPOINT else "NOT SET"
print(f"Azure OpenAI endpoint : {endpoint_display}")
print(f"Deployment            : {AZURE_OPENAI_DEPLOYMENT_NAME or 'NOT SET'}")
print(f"API version           : {AZURE_OPENAI_API_VERSION or 'NOT SET'}")
print(f"API key               : {'SET' if AZURE_OPENAI_API_KEY else 'NOT SET'}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 4 — Extractor (src/extractor.py)

# COMMAND ----------

import openpyxl
from datetime import datetime


def _safe(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def extract_checklist(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[CHECKLIST_TAB]
    result = {
        "preparer_name": _safe(ws[CHECKLIST_PREPARER_NAME_CELL].value),
        "preparer_date": _safe(ws[CHECKLIST_PREPARER_DATE_CELL].value),
        "reviewer_name": _safe(ws[CHECKLIST_REVIEWER_NAME_CELL].value),
        "reviewer_date": _safe(ws[CHECKLIST_REVIEWER_DATE_CELL].value),
    }
    signoffs = {}
    for attr, row in ATTRIBUTE_ROW_MAP.items():
        signoffs[attr] = {
            "row": row,
            "checked_by":   _safe(ws[f"{CHECKLIST_CHECKED_BY_COL}{row}"].value),
            "reviewed_by":  _safe(ws[f"{CHECKLIST_REVIEWED_BY_COL}{row}"].value),
            "annotation":   _safe(ws[f"{CHECKLIST_ANNOTATION_COL}{row}"].value),
            "ia_annotation":_safe(ws[f"{CHECKLIST_IA_ANNOTATION_COL}{row}"].value),
        }
    result["signoffs"] = signoffs
    all_rows = {}
    for row in range(CHECKLIST_APPROVAL_ROW_START, CHECKLIST_APPROVAL_ROW_END + 1):
        task    = _safe(ws[f"B{row}"].value)
        checked = _safe(ws[f"{CHECKLIST_CHECKED_BY_COL}{row}"].value)
        reviewed= _safe(ws[f"{CHECKLIST_REVIEWED_BY_COL}{row}"].value)
        if task or checked or reviewed:
            all_rows[row] = {"task": task, "checked_by": checked, "reviewed_by": reviewed}
    result["all_signoff_rows"] = all_rows
    comments = []
    for row in range(CHECKLIST_COMMENTS_ROW_START, CHECKLIST_COMMENTS_ROW_END + 1):
        val = _safe(ws[f"{CHECKLIST_COMMENTS_COL}{row}"].value)
        if val and str(val).strip() and str(val).strip() != "Reviewer/Approver Comments":
            comments.append({"row": row, "text": str(val).strip()})
    result["reviewer_comments"] = comments
    ia_annotations = []
    for row in range(CHECKLIST_APPROVAL_ROW_START, 36):
        val = _safe(ws[f"{CHECKLIST_IA_ANNOTATION_COL}{row}"].value)
        if val:
            ia_annotations.append({"row": row, "col": CHECKLIST_IA_ANNOTATION_COL, "text": str(val).strip()})
    result["ia_annotations"] = ia_annotations
    wb.close()
    return result


def extract_billing_summary(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[BILLING_SUMMARY_TAB]
    result = {
        "billing_rate_label":   _safe(ws[BS_BILLING_RATE_LABEL].value),
        "billing_rate_display": _safe(ws[BS_BILLING_RATE_DISPLAY].value),
        "billing_rate_decimal": _safe(ws[BS_BILLING_RATE_DECIMAL].value),
        "uk_gross":             _safe(ws[BS_UK_GROSS].value),
        "uk_less_cash":         _safe(ws[BS_UK_LESS_CASH].value),
        "uk_net":               _safe(ws[BS_UK_NET].value),
        "intl_gross":           _safe(ws[BS_INTL_GROSS].value),
        "intl_less_cash":       _safe(ws[BS_INTL_LESS_CASH].value),
        "intl_net":             _safe(ws[BS_INTL_NET].value),
        "hresdf_gross":         _safe(ws[BS_HRESDF_GROSS].value),
        "hresdf_less_income":   _safe(ws[BS_HRESDF_LESS_INCOME].value),
        "hresdf_net":           _safe(ws[BS_HRESDF_NET].value),
        "total_asset_value":    _safe(ws[BS_TOTAL_ASSET_VALUE].value),
        "quarterly_billings":   _safe(ws[BS_QUARTERLY_BILLINGS].value),
        "billing_check":        _safe(ws[BS_BILLING_CHECK].value),
    }
    disposals = []
    for row in range(BS_DISPOSAL_ROW_START, BS_DISPOSAL_ROW_END + 1):
        name = _safe(ws[f"{BS_DISPOSAL_NAME_COL}{row}"].value)
        if name and str(name).strip().lower() not in ("", "total", "other"):
            disposals.append({
                "row": row,
                "name": str(name).strip(),
                "days": _safe(ws[f"{BS_DISPOSAL_DAYS_COL}{row}"].value),
                "billing_amount": _safe(ws[f"{BS_DISPOSAL_BILLING_COL}{row}"].value),
            })
    result["disposals"] = disposals
    detail_box = {}
    for row in range(42, 50):
        p = _safe(ws[f"P{row}"].value)
        q = _safe(ws[f"Q{row}"].value)
        r = _safe(ws[f"R{row}"].value)
        if any(v is not None for v in [p, q, r]):
            detail_box[row] = {"P": p, "Q": q, "R": r}
    result["disposal_detail_box"] = detail_box
    rebates = []
    for row in range(BS_REBATE_ROW_START, BS_REBATE_ROW_END + 1):
        name   = _safe(ws[f"{BS_REBATE_NAME_COL}{row}"].value)
        amount = _safe(ws[f"{BS_REBATE_AMOUNT_COL}{row}"].value)
        if name and str(name).strip():
            rebates.append({
                "row": row,
                "name": str(name).strip(),
                "amount": amount,
                "rate_type": _safe(ws[f"{BS_REBATE_RATE_TYPE_COL}{row}"].value),
            })
    result["rebates"] = rebates
    wb.close()
    return result


def extract_formulas(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    ws = wb[BILLING_SUMMARY_TAB]
    formulas = {}
    for cell_ref in FORMULA_CELLS:
        val = ws[cell_ref].value
        formulas[cell_ref] = str(val) if val is not None else None
    wb.close()
    return formulas


def extract_reviewer_comments(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[CHECKLIST_TAB]
    entries = []
    for row in range(CHECKLIST_COMMENTS_ROW_START, CHECKLIST_COMMENTS_ROW_END + 1):
        val = _safe(ws[f"B{row}"].value)
        if val and str(val).strip() and str(val).strip() != "Reviewer/Approver Comments":
            entries.append({"row": row, "column": "B", "text": str(val).strip()})
    for row in range(CHECKLIST_APPROVAL_ROW_START, 36):
        val = _safe(ws[f"I{row}"].value)
        if val:
            entries.append({"row": row, "column": "I", "text": str(val).strip()})
    wb.close()
    return entries


def extract_all(file_path):
    return {
        "file_path":          file_path,
        "checklist":          extract_checklist(file_path),
        "billing_summary":    extract_billing_summary(file_path),
        "formulas":           extract_formulas(file_path),
        "reviewer_comments":  extract_reviewer_comments(file_path),
    }


print("Extractor functions defined.")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 5 — Attribute Checks (src/attribute_checks.py)
# MAGIC Layer 1 — deterministic Python. No LLM.

# COMMAND ----------

def _signoff(extracted_data, attr):
    return extracted_data["checklist"]["signoffs"].get(attr, {})


def _signed(signoff):
    return bool(signoff.get("checked_by")) and bool(signoff.get("reviewed_by"))


def check_attribute_a(extracted_data):
    formulas = extracted_data["formulas"]
    evidence = {}
    failures = []
    j35 = formulas.get("J35", "")
    evidence["J35_formula"] = j35
    if not j35 or not j35.startswith("="):
        failures.append("J35 is not a formula")
    elif "0.00285" not in j35 and "0.003" not in j35:
        failures.append(f"J35 formula does not reference expected billing rate: {j35}")
    j31 = formulas.get("J31", "")
    evidence["J31_formula"] = j31
    if not j31 or not j31.startswith("="):
        failures.append("J31 is not a formula")
    elif not all(x in j31 for x in ["J21", "J25", "J29"]):
        failures.append(f"J31 formula does not reference J21, J25, J29: {j31}")
    for cell in ["J21", "J25", "J29"]:
        val = formulas.get(cell, "")
        evidence[f"{cell}_formula"] = val
        if not val or not val.startswith("="):
            failures.append(f"{cell} is not a formula")
    signoff = _signoff(extracted_data, "A")
    evidence["signoff"] = signoff
    if not _signed(signoff):
        failures.append("Attribute A not signed off on Checklist")
    passed = len(failures) == 0
    return {"pass": passed, "evidence": evidence,
            "details": "All formulae verified and signed off." if passed else "; ".join(failures)}


def check_attribute_b(current_data, prior_data):
    current_bill = current_data["billing_summary"]["quarterly_billings"] or 0
    prior_bill   = prior_data["billing_summary"]["quarterly_billings"] or 0
    abs_movement = current_bill - prior_bill
    pct_movement = abs_movement / prior_bill if prior_bill else None
    exceeds_threshold = abs(pct_movement) > BILLING_MOVEMENT_THRESHOLD if pct_movement is not None else False
    evidence = {
        "current_quarterly_billings": current_bill,
        "prior_quarterly_billings":   prior_bill,
        "absolute_movement":          abs_movement,
        "percentage_movement":        round(pct_movement * 100, 2) if pct_movement is not None else None,
        "exceeds_5pct_threshold":     exceeds_threshold,
        "current_rate":               current_data["billing_summary"]["billing_rate_label"],
        "prior_rate":                 prior_data["billing_summary"]["billing_rate_label"],
        "reviewer_comments":          current_data["reviewer_comments"],
    }
    signoff = _signoff(current_data, "B")
    evidence["signoff"] = signoff
    passed = _signed(signoff)
    if passed:
        threshold_note = (
            f" Movement of {evidence['percentage_movement']}% exceeds 5% threshold — "
            "reviewer comments present." if exceeds_threshold else ""
        )
        details = (
            f"Reasonableness review signed off. Prior: £{prior_bill:,.2f}, "
            f"Current: £{current_bill:,.2f} ({evidence['percentage_movement']}% movement)."
            + threshold_note
        )
    else:
        details = "Attribute B not signed off on Checklist row 26."
    return {"pass": passed, "evidence": evidence, "details": details}


def check_attribute_c(extracted_data):
    bs = extracted_data["billing_summary"]
    evidence = {}
    failures = []
    checks = [
        ("UK",     bs.get("uk_gross"),     bs.get("uk_less_cash"),       bs.get("uk_net"),     "J21"),
        ("Intl",   bs.get("intl_gross"),   bs.get("intl_less_cash"),     bs.get("intl_net"),   "J25"),
        ("HRESDF", bs.get("hresdf_gross"), bs.get("hresdf_less_income"), bs.get("hresdf_net"), "J29"),
    ]
    net_sum = 0.0
    for label, gross, deduct, net, cell in checks:
        gross = gross or 0.0; deduct = deduct or 0.0; net = net or 0.0
        calculated = gross + deduct
        diff = abs(calculated - net)
        evidence[f"{label}_gross"] = gross; evidence[f"{label}_deduction"] = deduct
        evidence[f"{label}_net_per_sheet"] = net; evidence[f"{label}_calculated"] = calculated
        evidence[f"{label}_difference"] = diff
        if diff > FLOAT_TOLERANCE:
            failures.append(f"{label} ({cell}): calculated {calculated:,.2f} vs sheet {net:,.2f} — diff £{diff:,.2f}")
        net_sum += net
    j31 = bs.get("total_asset_value") or 0.0
    j31_diff = abs(j31 - net_sum)
    evidence["J31_total_asset_value"] = j31; evidence["J31_sum_of_nets"] = net_sum
    evidence["J31_difference"] = j31_diff
    if j31_diff > FLOAT_TOLERANCE:
        failures.append(f"J31: sheet {j31:,.2f} vs sum of nets {net_sum:,.2f} — diff £{j31_diff:,.2f}")
    signoff = _signoff(extracted_data, "C")
    evidence["signoff"] = signoff
    if not _signed(signoff):
        failures.append("Attribute C not signed off on Checklist")
    passed = len(failures) == 0
    return {"pass": passed, "evidence": evidence,
            "details": "All tie-throughs verified within £1 tolerance." if passed else "; ".join(failures)}


def check_attribute_d(current_data, prior_data):
    current_label   = current_data["billing_summary"]["billing_rate_label"]
    prior_label     = prior_data["billing_summary"]["billing_rate_label"]
    current_decimal = current_data["billing_summary"]["billing_rate_decimal"]
    prior_decimal   = prior_data["billing_summary"]["billing_rate_decimal"]
    billing_check   = current_data["billing_summary"]["billing_check"]
    ia_annotation   = current_data["checklist"]["signoffs"].get("D", {}).get("ia_annotation")
    reviewer_comments = current_data["reviewer_comments"]
    rate_changed = (current_label != prior_label) or (current_decimal != prior_decimal)
    evidence = {
        "current_rate_label":   current_label,
        "prior_rate_label":     prior_label,
        "current_rate_decimal": current_decimal,
        "prior_rate_decimal":   prior_decimal,
        "rate_changed":         rate_changed,
        "billing_check_result": billing_check,
        "ia_annotation":        ia_annotation,
        "reviewer_comments":    [c["text"] for c in reviewer_comments],
    }
    signoff = _signoff(current_data, "D")
    evidence["signoff"] = signoff
    if rate_changed:
        passed = False
        details = (
            f"EXCEPTION: Billing rate changed from {prior_label} to {current_label}. "
            f"Decimal: {prior_decimal} → {current_decimal}. "
            f"Billing check result: {billing_check}."
        )
        if ia_annotation:
            details += f" IA note: {ia_annotation}"
    else:
        passed = _signed(signoff)
        details = (
            f"Billing rate unchanged at {current_label} ({current_decimal}). Signed off."
            if passed else "Attribute D not signed off on Checklist."
        )
    return {"pass": passed, "evidence": evidence, "details": details}


def check_attribute_e(extracted_data):
    bs = extracted_data["billing_summary"]
    disposals = bs.get("disposals", [])
    evidence = {
        "disposals": disposals,
        "disposal_detail_box": bs.get("disposal_detail_box", {}),
        "main_billing_rate": bs.get("billing_rate_decimal"),
        "disposal_count": len(disposals),
    }
    signoff = _signoff(extracted_data, "E")
    evidence["signoff"] = signoff
    failures = []
    if not _signed(signoff):
        failures.append("Attribute E not signed off on Checklist row 30")
    passed = len(failures) == 0
    return {"pass": passed, "evidence": evidence,
            "details": f"Acquisitions/disposals reviewed. {len(disposals)} disposal(s). Signed off."
            if passed else "; ".join(failures)}


def check_attribute_f(extracted_data):
    rebates = extracted_data["billing_summary"].get("rebates", [])
    by_rate_type = {}
    for r in rebates:
        rt = r.get("rate_type") or "Unknown"
        by_rate_type.setdefault(rt, []).append({"name": r["name"], "amount": r["amount"]})
    evidence = {
        "rebates": rebates,
        "rebate_count": len(rebates),
        "by_rate_type": by_rate_type,
        "rate_types_found": list(by_rate_type.keys()),
    }
    signoff = _signoff(extracted_data, "F")
    evidence["signoff"] = signoff
    failures = []
    if not _signed(signoff):
        failures.append("Attribute F not signed off on Checklist row 31")
    passed = len(failures) == 0
    return {"pass": passed, "evidence": evidence,
            "details": f"Rebates reviewed. {len(rebates)} rebate(s) across "
            f"{len(by_rate_type)} rate type(s): {', '.join(by_rate_type.keys())}. Signed off."
            if passed else "; ".join(failures)}


def run_all_checks(q1_data, q3_data):
    results = {}
    for quarter, current, prior in [
        ("Q1", q1_data, q1_data),   # Q1 self-reference: rate unchanged → D passes
        ("Q3", q3_data, q1_data),   # Q3 vs Q1: rate changed → D fails
    ]:
        results[(quarter, "A")] = check_attribute_a(current)
        results[(quarter, "B")] = check_attribute_b(current, prior)
        results[(quarter, "C")] = check_attribute_c(current)
        results[(quarter, "D")] = check_attribute_d(current, prior)
        results[(quarter, "E")] = check_attribute_e(current)
        results[(quarter, "F")] = check_attribute_f(current)
    return results


print("Attribute check functions defined.")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 6 — LangChain Tools (src/tools.py)

# COMMAND ----------

from langchain.tools import tool


@tool
def read_checklist_signoffs(file_path: str) -> str:
    """Read the Checklist tab sign-offs from an AUM Billing workbook."""
    data = extract_checklist(file_path)
    lines = [
        f"Preparer: {data['preparer_name']} (date: {data['preparer_date']})",
        f"Reviewer: {data['reviewer_name']} (date: {data['reviewer_date']})",
        "", "Attribute sign-offs:",
    ]
    for attr, s in data["signoffs"].items():
        lines.append(
            f"  Attr {attr} (row {s['row']}): "
            f"checked_by={s['checked_by']}, reviewed_by={s['reviewed_by']}"
            + (f" | IA note: {s['ia_annotation']}" if s.get("ia_annotation") else "")
        )
    if data["ia_annotations"]:
        lines += ["", "IA annotations (column I):"]
        for a in data["ia_annotations"]:
            lines.append(f"  Row {a['row']}: {a['text']}")
    return "\n".join(lines)


@tool
def read_billing_rate(file_path: str) -> str:
    """Read the billing rate from the Billing Summary tab of an AUM Billing workbook."""
    bs = extract_billing_summary(file_path)
    return (
        f"Billing rate label (J33): {bs['billing_rate_label']}\n"
        f"Billing rate display (P17): {bs['billing_rate_display']}\n"
        f"Billing rate decimal (Q38): {bs['billing_rate_decimal']}\n"
        f"Quarterly billings (J35): £{bs['quarterly_billings']:,.2f}\n"
        f"Billing check (N49): {bs['billing_check']}"
    )


@tool
def read_reviewer_comments(file_path: str) -> str:
    """Read all reviewer comments from the Checklist tab. Use for exception investigation."""
    comments = extract_reviewer_comments(file_path)
    if not comments:
        return "No reviewer comments or IA annotations found."
    lines = []
    col_b = [c for c in comments if c["column"] == "B"]
    col_i = [c for c in comments if c["column"] == "I"]
    if col_b:
        lines.append("Reviewer/Approver comments (column B):")
        for c in col_b:
            lines.append(f"  Row {c['row']}: {c['text']}")
    if col_i:
        lines += ["", "IA annotations (column I):"]
        for c in col_i:
            lines.append(f"  Row {c['row']}: {c['text']}")
    return "\n".join(lines)


@tool
def read_disposals(file_path: str) -> str:
    """Read all acquisitions and disposals from the Billing Summary tab."""
    bs = extract_billing_summary(file_path)
    disposals = bs.get("disposals", [])
    if not disposals:
        return "No disposals found in Billing Summary."
    lines = [f"Disposals found ({len(disposals)}):"]
    for d in disposals:
        days = f", days={d['days']}" if d.get("days") else ""
        lines.append(f"  Row {d['row']}: {d['name']}{days}, billing adjustment=£{d['billing_amount']:,.2f}")
    lines.append(f"\nMain billing rate: {bs['billing_rate_label']} ({bs['billing_rate_decimal']})")
    return "\n".join(lines)


@tool
def read_rebates(file_path: str) -> str:
    """Read all rebates and ownership rates from the Billing Summary tab."""
    bs = extract_billing_summary(file_path)
    rebates = bs.get("rebates", [])
    if not rebates:
        return "No rebates found in Billing Summary."
    by_type = {}
    for r in rebates:
        rt = r.get("rate_type") or "Unknown"
        by_type.setdefault(rt, []).append(r)
    lines = [f"Rebates found ({len(rebates)}) grouped by rate type:"]
    for rt, items in by_type.items():
        lines.append(f"\n  Rate type: {rt}")
        for r in items:
            amt = r["amount"]
            lines.append(f"    {r['name']}: {'£{:,.2f}'.format(amt) if amt is not None else 'N/A'}")
    return "\n".join(lines)


@tool
def compare_billing_rates(q1_path: str, q3_path: str) -> str:
    """Compare billing rates between Q1 and Q3 workbooks to detect changes."""
    bs1 = extract_billing_summary(q1_path)
    bs3 = extract_billing_summary(q3_path)
    q1_label = bs1["billing_rate_label"]; q3_label = bs3["billing_rate_label"]
    q1_decimal = bs1["billing_rate_decimal"]; q3_decimal = bs3["billing_rate_decimal"]
    changed = (q1_label != q3_label) or (q1_decimal != q3_decimal)
    lines = [
        f"Q1 billing rate: {q1_label} ({q1_decimal})",
        f"Q3 billing rate: {q3_label} ({q3_decimal})",
        f"Rate changed: {'YES' if changed else 'NO'}",
    ]
    if changed:
        lines += [
            "", f"CHANGE DETECTED: {q1_label} → {q3_label}",
            f"Decimal change: {q1_decimal} → {q3_decimal}",
            f"Q1 billing check (N49): {bs1['billing_check']}",
            f"Q3 billing check (N49): {bs3['billing_check']}",
        ]
    return "\n".join(lines)


print("LangChain tools defined.")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 7 — LangChain Agent (src/agent.py)
# MAGIC Layer 2: direct LLM call for PASS narratives.
# MAGIC Layer 3: ReAct agent for EXCEPTION investigation.

# COMMAND ----------

import json
import time as _time
from langchain_openai import AzureChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage
from langgraph.prebuilt import create_react_agent

_llm = None


def _get_llm():
    global _llm
    if _llm is None:
        _llm = AzureChatOpenAI(
            azure_endpoint=os.environ["AZURE_OPENAI_ENDPOINT"],
            api_key=os.environ["AZURE_OPENAI_API_KEY"],
            azure_deployment=os.environ["AZURE_OPENAI_DEPLOYMENT_NAME"],
            api_version=os.environ["AZURE_OPENAI_API_VERSION"],
            temperature=0,
        )
    return _llm


NARRATIVE_SYSTEM_PROMPT = """You are an audit evidence narrator for Federated Hermes Limited.
Given verified audit findings, generate a professional narrative evidence summary for the DC-9 detective control.

Rules:
- Use ONLY the facts provided in the input JSON
- Do NOT infer, assume, or add any information not present in the input
- Include specific cell references and values where available
- NEVER use vague references like "row 25" or "row 30". Always reference cells
  using the compact format: WORKSHEET:<worksheet name>, CELL:<cell>. Examples:
    WORKSHEET:Checklist, CELL:F25   WORKSHEET:Billing Summary, CELL:J35
  The sign-off cell mapping is:
    Attribute A: WORKSHEET:Checklist, CELL:F25 (Checked By) and WORKSHEET:Checklist, CELL:G25 (Reviewed By)
    Attribute B: WORKSHEET:Checklist, CELL:F26 (Checked By) and WORKSHEET:Checklist, CELL:G26 (Reviewed By)
    Attribute C: WORKSHEET:Checklist, CELL:F28 (Checked By) and WORKSHEET:Checklist, CELL:G28 (Reviewed By)
    Attribute D: WORKSHEET:Checklist, CELL:F29 (Checked By) and WORKSHEET:Checklist, CELL:G29 (Reviewed By)
    Attribute E: WORKSHEET:Checklist, CELL:F30 (Checked By) and WORKSHEET:Checklist, CELL:G30 (Reviewed By)
    Attribute F: WORKSHEET:Checklist, CELL:F31 (Checked By) and WORKSHEET:Checklist, CELL:G31 (Reviewed By)
  For disposal rows, use "WORKSHEET:Billing Summary, CELL:B43", "WORKSHEET:Billing Summary, CELL:B44", etc.
  For rebate rows, use "WORKSHEET:Billing Summary, CELL:B55", "WORKSHEET:Billing Summary, CELL:B56", etc.
  For formula cells, use "WORKSHEET:Billing Summary, CELL:J35", etc.
- Include sign-off details (who prepared, who reviewed)
- Start with PASS ✓ followed by the attribute letter and description
- Always end with EXACTLY these two lines, substituting the actual names and dates from the input:
    Prepared & Checked by: {preparer_name} on {preparer_date}
    Reviewed by: {reviewer_name} on {reviewer_date}
  followed by a blank line and then:
    This confirms the sign-off on WORKSHEET:Checklist, CELL:F{row} and CELL:G{row}.
  where {row} is the attribute's sign-off row from the mapping above.
  Do NOT add any other sign-off text or annotations after this block.
- Keep it concise but factual — auditors need verifiable details
- Maximum 150 words"""

ATTRIBUTE_DESCRIPTIONS = {
    "A": "Checked formulae on Billing Summary tab",
    "B": "Reviewed for reasonableness against prior quarter",
    "C": "Checked figures tie through to backing sheets",
    "D": "Agreed no changes in billing rates since prior quarter",
    "E": "Agreed acquisitions/disposals to BTPS liquidity schedule",
    "F": "Checked rebates given at appropriate ownership rates",
}

EXCEPTION_SYSTEM_PROMPT = """You are an audit exception analyst for Federated Hermes Limited.
A deterministic check on the DC-9 detective control has FAILED. Your job is to:

1. Examine the failed check result provided
2. Use your tools to investigate — read reviewer comments, compare billing rates,
   look for explanations in the workbook
3. Generate an exception narrative that includes:
   - What changed (specific values from the check result)
   - Whether the change appears explained by reviewer comments or IMA references
   - What the auditor should verify independently
   - A recommendation (e.g., verify amended IMA document)

Start with EXCEPTION ✗ followed by the attribute letter and description.
Be specific about values.
NEVER use vague references like "row 29" or "row 30". Always reference cells
using the compact format: WORKSHEET:<worksheet name>, CELL:<cell>. The sign-off cell mapping is:
    Attribute A: WORKSHEET:Checklist, CELL:F25/G25
    Attribute B: WORKSHEET:Checklist, CELL:F26/G26
    Attribute C: WORKSHEET:Checklist, CELL:F28/G28
    Attribute D: WORKSHEET:Checklist, CELL:F29/G29
    Attribute E: WORKSHEET:Checklist, CELL:F30/G30
    Attribute F: WORKSHEET:Checklist, CELL:F31/G31
Always end the narrative with sign-off details including dates from the checklist:
    Prepared & Checked by: {preparer_name} on {preparer_date}
    Reviewed by: {reviewer_name} on {reviewer_date}
Maximum 200 words."""


def generate_narrative(attribute, quarter, check_result, extracted_data, max_retries=3):
    llm = _get_llm()
    attr_desc = ATTRIBUTE_DESCRIPTIONS.get(attribute, attribute)
    signoff_row = ATTRIBUTE_ROW_MAP.get(attribute)
    payload = {
        "quarter": quarter,
        "attribute": attribute,
        "attribute_description": attr_desc,
        "result": check_result,
        "preparer": extracted_data["checklist"].get("preparer_name"),
        "preparer_date": extracted_data["checklist"].get("preparer_date"),
        "reviewer": extracted_data["checklist"].get("reviewer_name"),
        "reviewer_date": extracted_data["checklist"].get("reviewer_date"),
        "signoff": extracted_data["checklist"]["signoffs"].get(attribute, {}),
        "signoff_cell_checked_by": f"WORKSHEET:Checklist, CELL:F{signoff_row}",
        "signoff_cell_reviewed_by": f"WORKSHEET:Checklist, CELL:G{signoff_row}",
        "billing_rate": extracted_data["billing_summary"].get("billing_rate_label"),
        "quarterly_billings": extracted_data["billing_summary"].get("quarterly_billings"),
    }
    user_msg = (
        f"Generate a narrative evidence summary for {quarter} Attribute {attribute}.\n\n"
        f"Input data:\n{json.dumps(payload, indent=2, default=str)}"
    )
    for attempt in range(max_retries):
        try:
            response = llm.invoke([
                SystemMessage(content=NARRATIVE_SYSTEM_PROMPT),
                HumanMessage(content=user_msg),
            ])
            return response.content
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"    Retry {attempt + 1}/{max_retries} for {quarter} Attr {attribute}: {e}")
                _time.sleep(2)
            else:
                return (
                    f"PASS ✓ Attribute {attribute} — {attr_desc}. "
                    f"[Narrative generation failed after {max_retries} attempts: {e}]"
                )


def investigate_exception(attribute, check_result, q1_data, q3_data, q1_path, q3_path, max_retries=3):
    llm = _get_llm()
    attr_desc = ATTRIBUTE_DESCRIPTIONS.get(attribute, attribute)
    tools = [read_checklist_signoffs, read_billing_rate, read_reviewer_comments,
             compare_billing_rates, read_disposals, read_rebates]
    agent = create_react_agent(llm, tools)

    current_data = q3_data if attribute == "D" else q1_data
    checklist = current_data.get("checklist", {})

    user_msg = (
        f"EXCEPTION DETECTED — {attribute}: {attr_desc}\n\n"
        f"Failed check result:\n{json.dumps(check_result, indent=2, default=str)}\n\n"
        f"Q1 file path: {q1_path}\n"
        f"Q3 file path: {q3_path}\n\n"
        f"Sign-off details for the narrative ending:\n"
        f"  Preparer: {checklist.get('preparer_name')} on {checklist.get('preparer_date')}\n"
        f"  Reviewer: {checklist.get('reviewer_name')} on {checklist.get('reviewer_date')}\n\n"
        f"Investigate the exception using your tools. Read reviewer comments and "
        f"compare billing rates. Then generate a professional exception narrative."
    )
    for attempt in range(max_retries):
        try:
            response = agent.invoke({
                "messages": [
                    SystemMessage(content=EXCEPTION_SYSTEM_PROMPT),
                    HumanMessage(content=user_msg),
                ]
            })
            return response["messages"][-1].content
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"    Retry {attempt + 1}/{max_retries} for exception Attr {attribute}: {e}")
                _time.sleep(2)
            else:
                ev = check_result.get("evidence", {})
                return (
                    f"EXCEPTION ✗ Attribute {attribute} — {attr_desc}. "
                    f"Prior rate: {ev.get('prior_rate_label')}, "
                    f"Current rate: {ev.get('current_rate_label')}. "
                    f"Billing check: {ev.get('billing_check_result')}. "
                    f"IA note: {ev.get('ia_annotation', 'None')}. "
                    f"[Agent investigation failed: {e}]"
                )


def generate_all_narratives(q1_data, q3_data, check_results, q1_path, q3_path):
    narratives = {}
    total = len(check_results)
    for i, ((quarter, attr), result) in enumerate(check_results.items(), 1):
        extracted = q1_data if quarter == "Q1" else q3_data
        layer = "Layer 2" if result["pass"] else "Layer 3 (agent)"
        print(f"  [{i}/{total}] {quarter} Attr {attr} "
              f"({'PASS' if result['pass'] else 'EXCEPTION'}) — {layer}...", end=" ", flush=True)
        t0 = _time.time()
        if result["pass"]:
            narrative = generate_narrative(attr, quarter, result, extracted)
        else:
            narrative = investigate_exception(attr, result, q1_data, q3_data, q1_path, q3_path)
        narratives[(quarter, attr)] = narrative
        print(f"done ({_time.time() - t0:.1f}s)")
    return narratives


print("Agent functions defined.")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 8 — Excel Writer (src/excel_writer.py)

# COMMAND ----------

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime as _dt

_HEADER_FILL              = PatternFill(fill_type="solid", fgColor="1F3864")
_HEADER_FONT              = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
_NARRATIVE_FILL           = PatternFill(fill_type="solid", fgColor="DEEAF1")
_NARRATIVE_FONT           = Font(name="Calibri", size=9)
_NARRATIVE_FONT_EXCEPTION = Font(name="Calibri", size=9, color="FF0000")
_NARRATIVE_ALIGNMENT      = Alignment(wrap_text=True, vertical="top")
_NARRATIVE_ROW_HEIGHT     = 200
_THIN_SIDE   = Side(style="thin", color="000000")
_THIN_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def write_checkmarks(ws, check_results):
    for (quarter, attr), result in check_results.items():
        col = TOC_CHECKMARK_COLS.get(attr)
        if col is None:
            continue
        row = TOC_Q1_RESULTS_ROW if quarter == "Q1" else TOC_Q3_RESULTS_ROW
        cell = ws[f"{col}{row}"]
        if result["pass"]:
            cell.value = CHECKMARK_VALUE
            cell.font = Font(name=CHECKMARK_FONT_NAME, size=CHECKMARK_FONT_SIZE)
        else:
            cell.value = EXCEPTION_VALUE
            cell.font = Font(name=EXCEPTION_FONT_NAME, size=EXCEPTION_FONT_SIZE, bold=True, color="FF0000")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_narrative_headers(ws):
    for col, header in [
        (TOC_Q1_NARRATIVE_COL, TOC_Q1_NARRATIVE_HEADER),
        (TOC_Q3_NARRATIVE_COL, TOC_Q3_NARRATIVE_HEADER),
    ]:
        cell = ws[f"{col}{TOC_NARRATIVE_HEADER_ROW}"]
        cell.value = header
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def write_narratives(ws, narratives):
    for (quarter, attr), narrative in narratives.items():
        row = TOC_ATTR_ROW_MAP.get(attr)
        if row is None:
            continue
        col = TOC_Q1_NARRATIVE_COL if quarter == "Q1" else TOC_Q3_NARRATIVE_COL
        cell = ws[f"{col}{row}"]
        cell.value = narrative
        cell.font = _NARRATIVE_FONT_EXCEPTION if narrative.startswith("EXCEPTION") else _NARRATIVE_FONT
        cell.fill = _NARRATIVE_FILL
        cell.alignment = _NARRATIVE_ALIGNMENT
        cell.border = _THIN_BORDER
    for row in range(TOC_NARRATIVE_ROW_START, TOC_NARRATIVE_ROW_END + 1):
        ws.row_dimensions[row].height = _NARRATIVE_ROW_HEIGHT


def write_conclusion(ws, check_results):
    any_exception = any(not r["pass"] for r in check_results.values())
    cell = ws[TOC_EXCEPTIONS_NOTED_CELL]
    cell.value = "Yes" if any_exception else "No"
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000") if any_exception \
        else Font(name="Calibri", size=11)


def write_ai_metadata(ws):
    timestamp = _dt.now().strftime("%Y-%m-%d %H:%M")
    lines = [
        f"AI-generated narratives — {timestamp}",
        "Model: Azure OpenAI gpt-4o-mini (LangChain / LangGraph)",
        "Layer 2: Direct LLM call (PASS narratives)",
        "Layer 3: ReAct agent investigation (EXCEPTION narratives)",
    ]
    for i, text in enumerate(lines):
        cell = ws[f"P{50 + i}"]
        cell.value = text
        cell.font = Font(name="Calibri", size=8, italic=True, color="595959")
        cell.alignment = Alignment(wrap_text=True)


def write_toc(toc_path, output_path, check_results, narratives):
    os.makedirs(DATA_OUTPUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(toc_path)
    if TOC_DC9_TAB not in wb.sheetnames:
        raise ValueError(f"Tab '{TOC_DC9_TAB}' not found. Available: {wb.sheetnames}")
    ws = wb[TOC_DC9_TAB]
    write_narrative_headers(ws)
    write_checkmarks(ws, check_results)
    write_narratives(ws, narratives)
    write_conclusion(ws, check_results)
    write_ai_metadata(ws)
    wb.save(output_path)
    return output_path


print("Excel writer functions defined.")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 8b — File Arrival Guard Clause
# MAGIC
# MAGIC Ensures **both** Q1 and Q3 workbooks are present before the pipeline runs.
# MAGIC
# MAGIC The Databricks File Arrival Trigger fires when a file lands in
# MAGIC `data/input/`. Because Storage Explorer uploads Q1 and Q3 in parallel
# MAGIC there can be a brief window (typically 1–10 seconds) where one file
# MAGIC is visible and the other is still syncing from ADLS to the UC Volume.
# MAGIC
# MAGIC This cell polls for up to 90 seconds, waiting for both files to
# MAGIC appear, and hard-fails if the TOC template is missing. Without this
# MAGIC guard, a racing trigger can produce a half-populated run.

# COMMAND ----------

import os
import time as _time_guard

GUARD_TIMEOUT_SECONDS = 90
GUARD_POLL_INTERVAL   = 5

print("=" * 60)
print("  File Arrival Guard — waiting for Q1 + Q3 + TOC")
print("=" * 60)
print(f"  Q1_FILE  : {Q1_FILE}")
print(f"  Q3_FILE  : {Q3_FILE}")
print(f"  TOC_FILE : {TOC_FILE}")
print()

# TOC template must be pre-staged (day before demo). Fail loudly if missing.
if not os.path.exists(TOC_FILE):
    raise FileNotFoundError(
        f"TOC template missing: {TOC_FILE}\n"
        f"  → Upload 4_06_2025_Revenue _TOC.xlsx to data/toc/ before running."
    )
print(f"  \u2713 TOC template found")

# Wait for Q1 + Q3 to both be present
deadline = _time_guard.time() + GUARD_TIMEOUT_SECONDS
last_status = None
while _time_guard.time() < deadline:
    q1_ok = os.path.exists(Q1_FILE)
    q3_ok = os.path.exists(Q3_FILE)
    status = (q1_ok, q3_ok)

    if q1_ok and q3_ok:
        print(f"  \u2713 Q1 found")
        print(f"  \u2713 Q3 found")
        print()
        print("  All required files present \u2014 proceeding with pipeline.")
        print("=" * 60)
        break

    if status != last_status:
        waiting = []
        if not q1_ok: waiting.append("Q1")
        if not q3_ok: waiting.append("Q3")
        remaining = int(deadline - _time_guard.time())
        print(f"  \u2026 waiting for {', '.join(waiting)}  ({remaining}s remaining)")
        last_status = status

    _time_guard.sleep(GUARD_POLL_INTERVAL)
else:
    raise FileNotFoundError(
        f"Timed out after {GUARD_TIMEOUT_SECONDS}s waiting for both workbooks.\n"
        f"  Q1 exists: {os.path.exists(Q1_FILE)}\n"
        f"  Q3 exists: {os.path.exists(Q3_FILE)}\n"
        f"  → Drop both Q1 and Q3 into data/input/ simultaneously."
    )

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 9 — Run Pipeline
# MAGIC Execute all four steps end-to-end.

# COMMAND ----------

import time

def _banner(text):
    print(); print("=" * 60); print(f"  {text}"); print("=" * 60)

def _step(n, label):
    print(); print(f"Step {n}: {label}"); print("-" * 50)

def _elapsed(seconds):
    return f"{seconds:.1f}s" if seconds < 60 else f"{int(seconds // 60)}m {seconds % 60:.0f}s"


pipeline_start = time.time()

_banner("DC-9 Automated Audit Control Testing")
print(f"  Q1  : {Q1_FILE.split('/')[-1]}")
print(f"  Q3  : {Q3_FILE.split('/')[-1]}")
print(f"  TOC : {TOC_FILE.split('/')[-1]}")
print(f"  Out : {OUTPUT_FILE.split('/')[-1]}")

# Step 1 — Extract
_step(1, "Extracting data from source workbooks...")
t0 = time.time()
q1_data = extract_all(Q1_FILE)
q3_data = extract_all(Q3_FILE)
print(f"  Q1: billing rate={q1_data['billing_summary']['billing_rate_label']}, "
      f"quarterly billings=£{q1_data['billing_summary']['quarterly_billings']:,.2f}")
print(f"  Q3: billing rate={q3_data['billing_summary']['billing_rate_label']}, "
      f"quarterly billings=£{q3_data['billing_summary']['quarterly_billings']:,.2f}")
print(f"  Done ({_elapsed(time.time() - t0)})")

# Step 2 — Attribute checks
_step(2, "Running attribute checks (deterministic)...")
t0 = time.time()
check_results = run_all_checks(q1_data, q3_data)
passes    = sum(1 for r in check_results.values() if r["pass"])
exceptions= sum(1 for r in check_results.values() if not r["pass"])
for (quarter, attr), result in sorted(check_results.items()):
    print(f"  {quarter} Attr {attr}: {'PASS' if result['pass'] else '*** EXCEPTION ***'}")
print(f"\n  Results: {passes} PASS, {exceptions} EXCEPTION")
print(f"  Done ({_elapsed(time.time() - t0)})")

# Step 3 — Narratives
_step(3, "Generating evidence narratives (Azure OpenAI)...")
t0 = time.time()
try:
    narratives = generate_all_narratives(q1_data, q3_data, check_results, Q1_FILE, Q3_FILE)
    print(f"  Done ({_elapsed(time.time() - t0)})")
except Exception as e:
    print(f"  WARNING: {e} — falling back to deterministic summaries.")
    narratives = {
        (q, a): (
            f"PASS ✓ Attribute {a} — {q}. {r.get('details', '')}"
            if r["pass"] else
            f"EXCEPTION ✗ Attribute {a} — {q}. {r.get('details', '')}"
        )
        for (q, a), r in check_results.items()
    }
    print(f"  Done with fallback ({_elapsed(time.time() - t0)})")

# Step 4 — Write
_step(4, "Writing results to TOC workbook...")
t0 = time.time()
output_path = write_toc(TOC_FILE, OUTPUT_FILE, check_results, narratives)
print(f"  Output written: {output_path}")
print(f"  Done ({_elapsed(time.time() - t0)})")

# Summary
total_time = time.time() - pipeline_start
_banner("Pipeline Complete")
print(f"  Total time : {_elapsed(total_time)}")
print(f"  Checks     : {passes} PASS / {exceptions} EXCEPTION")
print(f"  Narratives : {len(narratives)} generated")
print(f"  Output     : {output_path}")

# COMMAND ----------

# MAGIC %md
# MAGIC ## Cell 10 — Download output (Databricks only)
# MAGIC Run this cell to make the output file available for download from the Databricks UI.

# COMMAND ----------

# Databricks: display download link for the completed TOC workbook
# Uncomment when running on cluster:

# dbutils.fs.cp(
#     f"file:{OUTPUT_FILE}",
#     f"dbfs:/FileStore/dc9_output/{OUTPUT_FILENAME}"
# )
# displayHTML(
#     f'<a href="/files/dc9_output/{OUTPUT_FILENAME}" target="_blank">'
#     f'Download {OUTPUT_FILENAME}</a>'
# )

print("To download the output on Databricks: uncomment the dbutils.fs.cp block above.")
print(f"Output file location: {OUTPUT_FILE}")
